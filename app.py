import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import numpy as np
from io import BytesIO
import os
from PIL import Image
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import Tuple, List, Optional

# ==========================================
# 1. CONFIGURACIÓN Y CONSTANTES
# ==========================================

class Config:
    PAGE_TITLE = "Inducor - Diagnóstico IEEE 400.2"
    LOGO_PATH = "logo_inducor.png"
    
    FAN_COLORS = [
        "#FFFF00", "#0D47A1", "#2979FF", "#00B0FF", "#00C853", 
        "#64DD17", "#AEEA00", "#FFD600", "#FF6D00", "#D50000", "#212121"
    ]
    
    COLOR_GRID_LINES = "#E0E0E0"
    COLOR_DASHED_LINE = "#81C784"
    COLOR_TEXT_MAIN = "black"
    COLOR_AREA_FILL = "rgba(255, 165, 0, 0.4)"
    COLOR_AREA_BORDER = "orange"

st.set_page_config(page_title=Config.PAGE_TITLE, layout="wide")


# ==========================================
# 2. LÓGICA DE NEGOCIO
# ==========================================

def calculate_theoretical_voltages(u_nominal: float) -> Tuple[float, float, float]:
    if u_nominal == 0:
        return 0.0, 0.0, 0.0
    uo_math = round(u_nominal / (3**0.5), 2)
    return round(uo_math * 0.5, 2), uo_math, round(uo_math * 1.5, 2)

def generate_hybrid_x_axis(voltage_values: List[float], show_labels: bool = False) -> Tuple[List[float], List[str]]:
    if not voltage_values or len(voltage_values) < 3:
        return [], []
    
    v05, v10, v15 = voltage_values[0], voltage_values[1], voltage_values[2]
    
    min_int = int(np.floor(v05))
    max_int = int(np.ceil(v15))
    ticks_integers = list(range(min_int, max_int + 2))
    
    ticks_filtered = []
    for t in ticks_integers:
        is_conflicting = False
        for v in voltage_values:
            if abs(t - v) < 0.6: 
                is_conflicting = True
                break
        if not is_conflicting:
            ticks_filtered.append(t)

    final_ticks = sorted(list(set(ticks_filtered + voltage_values)))
    
    tick_labels = []
    for t in final_ticks:
        is_key_value = False
        label_text = ""

        if abs(t - v05) < 0.01:
            label_text = "0.5 U₀"
            is_key_value = True
        elif abs(t - v10) < 0.01:
            label_text = "1.0 U₀"
            is_key_value = True
        elif abs(t - v15) < 0.01:
            label_text = "1.5 U₀"
            is_key_value = True

        if is_key_value and show_labels:
            tick_labels.append(f"<b>{t:.2f}</b><br>{label_text}")
        elif is_key_value:
            tick_labels.append(f"<b>{t:.2f}</b>")
        else:
            tick_labels.append(str(int(t)))
            
    return final_ticks, tick_labels

def calculate_fan_intersections(u0_val: float, voltage_points: List[float]) -> pd.DataFrame:
    c_values = np.arange(u0_val, u0_val + 5.1, 0.5)
    data = []
    for c in c_values:
        a = max(0.0, c - 5.0)
        data.append({
            "0,5·U0": f"{a:.2f}",
            "U0": f"{u0_val:.2f}",
            "1,5·U0": f"{c:.2f}"
        })
    return pd.DataFrame(data)


# ==========================================
# 3. MOTOR DE GRÁFICOS (PLOTLY)
# ==========================================

class ChartBuilder:

    @staticmethod
    def _add_logo(fig: go.Figure, x_pos: float, y_pos: float, anchor: str):
        if os.path.exists(Config.LOGO_PATH):
            try:
                img = Image.open(Config.LOGO_PATH)
                fig.add_layout_image(dict(
                    source=img, xref="paper", yref="paper",
                    x=x_pos, y=y_pos,
                    sizex=0.20 if anchor == "left" else 0.25,
                    sizey=0.20 if anchor == "left" else 0.25,
                    xanchor=anchor, yanchor="bottom"
                ))
            except Exception:
                pass

    @staticmethod
    def _add_vertical_lines_with_labels(fig: go.Figure, voltages: List[float]):
        labels = ["0,5 U₀", "U₀", "1,5 U₀"]
        for i, voltage in enumerate(voltages):
            fig.add_vline(x=voltage, line_dash="dash", line_color=Config.COLOR_DASHED_LINE, line_width=2)
            fig.add_annotation(
                x=voltage, y=1.0, yref="paper", 
                text=f"<b>{labels[i]}</b>",
                showarrow=False, yshift=20, 
                font=dict(size=14, color="black", family="Arial Black"),
                bgcolor="rgba(255,255,255,0.7)"
            )

    @staticmethod
    def _apply_layout(fig: go.Figure, title: str, x_ticks: list, x_labels: list, max_y: float, strict_x_range: bool = False, x_bounds: list = None):
        xaxis_config = dict(
            title=dict(text="<b>NIVELES DE TENSIÓN (kV)</b>", font=dict(color=Config.COLOR_TEXT_MAIN, size=14)),
            tickmode='array', tickvals=x_ticks, ticktext=x_labels, tickangle=0,
            gridcolor='rgba(0,0,0,0)', linecolor=Config.COLOR_TEXT_MAIN, mirror=True,
            tickfont=dict(size=11, color=Config.COLOR_TEXT_MAIN, family="Arial")
        )

        if strict_x_range and x_bounds:
            xaxis_config['range'] = x_bounds

        fig.update_layout(
            title=dict(
                text=f"<b>{title}</b>", x=0.5, xanchor='center', 
                font=dict(size=18, color=Config.COLOR_TEXT_MAIN, family="Arial Black")
            ),
            xaxis=xaxis_config,
            yaxis=dict(
                title=dict(text="<b>TG (δ) 1·10⁻³</b>", font=dict(color=Config.COLOR_TEXT_MAIN, size=14)),
                range=[0, max_y], dtick=1,
                gridcolor=Config.COLOR_GRID_LINES, linecolor=Config.COLOR_TEXT_MAIN, mirror=True,
                tickfont=dict(size=12, color=Config.COLOR_TEXT_MAIN, family="Arial")
            ),
            plot_bgcolor='white', paper_bgcolor='white',
            height=650, margin=dict(l=80, r=80, t=100, b=220),
            showlegend=False
        )

    @staticmethod
    def create_fan_chart(voltage_list: List[float], u0_val: float, u_nominal: float) -> go.Figure:
        fig = go.Figure()
        x_start, x_mid, x_end = voltage_list
        
        fan_steps = np.arange(u0_val, u0_val + 5.5, 0.5)
        for i, v15_final in enumerate(reversed(fan_steps)):
            v05_start = max(0.0, v15_final - 5.0)
            idx_color = min(i, len(Config.FAN_COLORS) - 1)
            
            fig.add_trace(go.Scatter(
                x=[x_start, x_mid, x_end],
                y=[v05_start, u0_val, v15_final],
                mode='lines',
                line=dict(color=Config.FAN_COLORS[idx_color], width=2),
                hovertemplate="<b>Tensión:</b> %{x:.2f} kV<br><b>Tg δ:</b> %{y:.2f} x10⁻³<extra></extra>",
                showlegend=False
            ))

        ChartBuilder._add_vertical_lines_with_labels(fig, voltage_list)

        fig.add_annotation(
            x=x_start, y=u0_val, text=f"<b>{u0_val:.2f}</b>",
            showarrow=True, arrowhead=2, arrowsize=1.5, ax=60, ay=-30,
            font=dict(size=16, color="black"), arrowcolor="#424242"
        )
        fig.add_annotation(
            x=x_mid, y=u0_val, text=f"<b>{u0_val:.2f}</b>",
            showarrow=True, arrowhead=2, arrowsize=1.5, ax=-60, ay=-30,
            font=dict(size=16, color="black"), arrowcolor="#424242"
        )
        tope_teorico = u0_val + 5.0
        fig.add_annotation(
            x=x_end, y=tope_teorico, text=f"<b>{tope_teorico:.2f}</b>",
            xanchor="right", showarrow=True, arrowhead=2, arrowsize=1.5, ax=-60, ay=-30,
            font=dict(size=16, color="black"), arrowcolor="#424242"
        )

        ChartBuilder._add_logo(fig, 1.0, -0.38, "right")
        
        ticks_v, ticks_l = generate_hybrid_x_axis(voltage_list, show_labels=False)
        max_y = int(u0_val + 6)
        title = f"RESULTADO DE ENSAYO DE TG (δ)<br>CABLE DE MEDIA TENSIÓN ({u_nominal}kV)"
        
        ChartBuilder._apply_layout(fig, title, ticks_v, ticks_l, max_y, strict_x_range=False)
        return fig

    @staticmethod
    def create_safety_area_chart(voltage_list: List[float], u0_val: float, u_nominal: float) -> go.Figure:
        fig = go.Figure()
        
        x_start, x_mid, x_end = voltage_list
        dynamic_ceiling = u0_val + 5.0
        area_y_points = [u0_val, u0_val, dynamic_ceiling]
        
        fig.add_trace(go.Scatter(
            x=[x_start, x_mid, x_end],
            y=area_y_points,
            mode='lines', line=dict(width=0),
            fill='tozeroy', fillcolor=Config.COLOR_AREA_FILL,
            name='ZONA SEGURA',
            hoverinfo='skip' 
        ))

        fig.add_trace(go.Scatter(
            x=[x_start, x_mid, x_end],
            y=area_y_points,
            mode='lines', line=dict(color=Config.COLOR_AREA_BORDER, width=3, dash='solid'),
            showlegend=False,
            hoverinfo='skip'
        ))

        fig.add_annotation(
            x=x_start, y=u0_val, 
            text=f"<b>{u0_val:.2f} [LIMIT]</b>",
            xanchor="left", yanchor="bottom", showarrow=False,
            font=dict(color="black", size=14, family="Arial Black"), yshift=5, xshift=5
        )
        
        fig.add_annotation(
            x=x_mid, y=u0_val, 
            text=f"<b>{u0_val:.2f} [LIMIT]</b>",
            xanchor="center", yanchor="bottom", showarrow=False,
            font=dict(color="black", size=14, family="Arial Black"), yshift=5
        )

        fig.add_annotation(
            x=x_end, y=dynamic_ceiling, 
            text=f"<b>{dynamic_ceiling:.2f} [LIMIT]</b>",
            xanchor="right", yanchor="bottom", showarrow=False,
            font=dict(color="black", size=14, family="Arial Black"), yshift=5, xshift=-5
        )

        x_center = (x_mid + x_end) / 2 
        y_center = (u0_val + 2.5) / 2
        fig.add_annotation(
            x=x_center, y=y_center,
            text="<b>IEEE400.2 MERIT AREA:<br>NO ACTION REQUIRED</b>",
            showarrow=False, font=dict(size=14, color="black", family="Arial Black"),
            align="center", bgcolor="rgba(255, 255, 255, 0.5)", bordercolor="black", borderwidth=1
        )

        ChartBuilder._add_logo(fig, 1.0, -0.38, "right")
        
        # Etiquetas EXTRA activadas solo aquí
        ticks_v, ticks_l = generate_hybrid_x_axis(voltage_list, show_labels=True)
        max_view_y = max(10, dynamic_ceiling + 1)
        title = f"TAN DELTA ON {u_nominal}KV CLASS CABLE - NO ACTION GRAPHIC AREA FOR:<br>U₀≤4·10⁻³ & [Tgδ@1.5U₀ – Tgδ@0.5U₀]≤1.5x10⁻³"
        
        x_bounds = [x_start, x_end]
        ChartBuilder._apply_layout(fig, title, ticks_v, ticks_l, max_view_y, strict_x_range=True, x_bounds=x_bounds)
        return fig


# ==========================================
# 4. GENERACIÓN DE REPORTES (EXCEL)
# ==========================================

def generate_excel_report(fig_fan: go.Figure, fig_area: go.Figure, u0_val: float, voltages: list) -> Optional[bytes]:
    try:
        output = BytesIO()
        img_fan = fig_fan.to_image(format="png", width=1400, height=800, scale=2)
        img_area = fig_area.to_image(format="png", width=1400, height=800, scale=2)
        df_data = calculate_fan_intersections(u0_val, voltages)

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_data.to_excel(writer, index=False, sheet_name='Reporte', startrow=1, startcol=1)
            ws = writer.sheets['Reporte']
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center')

            for cell in ws[2]: 
                if cell.column_letter in ['B', 'C', 'D']:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border

            for row in ws.iter_rows(min_row=3, max_row=len(df_data)+2, min_col=2, max_col=4):
                for cell in row:
                    cell.alignment = center_align
                    cell.border = border

            for col in ['B', 'C', 'D']:
                ws.column_dimensions[col].width = 18

            fila_inicio_graficos = len(df_data) + 5
            ws.add_image(OpenpyxlImage(BytesIO(img_fan)), f'B{fila_inicio_graficos}')
            ws.add_image(OpenpyxlImage(BytesIO(img_area)), f'B{fila_inicio_graficos + 45}')
            
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error generando reporte: {str(e)}")
        return None


# ==========================================
# 5. INTERFAZ DE USUARIO (MAIN)
# ==========================================

def main():
    if os.path.exists(Config.LOGO_PATH):
        try:
            st.sidebar.image(Image.open(Config.LOGO_PATH), width=220)
        except Exception:
            st.sidebar.title("INDUCOR")
    else:
        st.sidebar.title("INDUCOR")

    st.sidebar.header("Parámetros de Ensayo")
    
    u_linea = st.sidebar.number_input(
        "Tensión Nominal (U) [kV]", 
        min_value=0.0, value=0.0, step=0.1
    )

    v05_def, u0_def, v15_def = calculate_theoretical_voltages(u_linea)

    with st.sidebar.expander("Ajuste Fino de Ejes (kV)", expanded=True):
        st.caption("Modifique estos valores para ajustar el eje X a su protocolo:")
        kv_start = st.number_input("Tensión Inicial (0.5 Uo)", value=v05_def, step=0.1, format="%.2f")
        kv_mid = st.number_input("Tensión Central (Uo)", value=u0_def, step=0.1, format="%.2f")
        kv_end = st.number_input("Tensión Final (1.5 Uo)", value=v15_def, step=0.1, format="%.2f")

    u0_measured = st.sidebar.number_input(
        "Tan Delta en Uo (1.0) [Referencia Base]", 
        min_value=0.0, max_value=4.0, value=0.0, step=0.01,
        help="Valor base para construir las curvas de referencia."
    )

    if u_linea > 0:
        voltage_points = [kv_start, kv_mid, kv_end]

        try:
            tab_fan, tab_area = st.tabs(["Tendencia ", "Área No Action"])
            
            with tab_fan:
                fig_fan = ChartBuilder.create_fan_chart(voltage_points, u0_measured, u_linea)
                st.plotly_chart(fig_fan, use_container_width=True)
                
            with tab_area:
                fig_area = ChartBuilder.create_safety_area_chart(voltage_points, u0_measured, u_linea)
                st.plotly_chart(fig_area, use_container_width=True)

            st.markdown("---")
            if st.sidebar.button("Generar Informe de Ingeniería"):
                with st.spinner("Generando reporte..."):
                    excel_data = generate_excel_report(fig_fan, fig_area, u0_measured, voltage_points)
                    
                    if excel_data:
                        st.sidebar.success("¡Informe generado con éxito!")
                        st.sidebar.download_button(
                            label="Descargar Reporte (.xlsx)",
                            data=excel_data,
                            file_name=f"Informe_Inducor_{u_linea}kV.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

        except Exception as e:
            st.error(f"Se produjo un error al renderizar: {e}")
            st.exception(e)
    else:
        st.info("Hola, Ingrese la Tensión Nominal para comenzar.")

if __name__ == "__main__":
    main()